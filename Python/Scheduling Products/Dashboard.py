import psycopg2
import dash
from dash import dcc, html, Input, Output, dash_table
from dash.dependencies import Input, Output,State, MATCH, ALL
from datetime import datetime, timedelta
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import pandas as pd
import dash_bootstrap_components as dbc  # Import dash-bootstrap-components
import subprocess
import os
from dash.exceptions import PreventUpdate
import threading
import time
import json
import queue
import matplotlib.pyplot as plt
from threading import Timer
from openpyxl import load_workbook
import sys


allocation_process=None

def launch_dashboard():

    BOM=None
    Dash_time=None
    global flag1
    
    full_file_path=r'Data Sample.xlsx'
    # Function to fetch data from the database
    def fetch_data(full_file_path, sheet_name):
        
        try:
            data = pd.read_excel(full_file_path, sheet_name=sheet_name)
            
            return data
        except FileNotFoundError:
            print(f"Error: The file at {full_file_path} was not found.")
            return None
        except pd.errors.EmptyDataError:
            print(f"Error: The file at {full_file_path} is empty.")
            return None
        except pd.errors.ParserError:
            print(f"Error: The file at {full_file_path} does not appear to be a valid Excel file.")
            return None
        except Exception as e:
            print(f"Error fetching data from {sheet_name}: {e}")
            return None


    # Function to fetch data from the Excel file
    def fetch_data_table(full_file_path, sheet_name, usecols=None):
        try:
            data = pd.read_excel(full_file_path, sheet_name=sheet_name, usecols=usecols)
            
            return data
        except Exception as e:
            print(f"Error fetching data from {sheet_name}: {e}")
            return None
        
    # Define button style
    button_style = {
        'margin': '10px',
        'padding': '15px 30px',
        'font-size': '16px',
        'font-weight': 'bold',
        'border-radius': '8px',
        'background-color': '#3498db',
        'color': 'white',
        'border': 'none',
        'cursor': 'pointer',
        'transition': 'background-color 0.3s ease',
    }
    
    app = dash.Dash(__name__, external_stylesheets=[dbc.themes.SLATE])
    app.config.suppress_callback_exceptions = True

    # Define your CSS styles (assuming you have the CSS file as described)
    app.css.append_css({'external_url': '/assets/styles.css'})
    # Define layout
    app.layout = dbc.Container(
       
        style={'textAlign': 'left', 'width': '100%', 'margin': 'auto'},
        children=[
            html.Div(style={'height': '50px'}),
            html.H1('Dashboard - BOM Analysis', style={'textAlign': 'center', 'marginBottom': '30px'}),

            dcc.Interval(
                id='interval-component-script',
                interval=1000,  # Update script control every second
                n_intervals=0,
                disabled=True  # Start disabled
            ),

            dbc.Row([
                dbc.Col([
                    html.Button('Read from Spreadsheet', id='read-button', n_clicks=0, style=button_style),
                    html.Button('Start', id='initialise-button', n_clicks=0, style=button_style),
                    html.Button('Reschedule', id='start-button', n_clicks=0, style=button_style),
                    html.Button('Pause', id='stop-button', n_clicks=0, style=button_style),
                    html.Button('Reset', id='reset-button', n_clicks=0, style=button_style),
                    html.Div(id='start-message', style={'marginLeft': '20px', 'color': 'green'})
                ], width=15, style={'textAlign': 'Right', 'margin': 'auto'})
            ]),
            
            # Modals for feedback messages
            dbc.Modal(
                [
                    dbc.ModalHeader("Operation Status"),
                    dbc.ModalBody(id='read-modal-body'),
                    dbc.ModalFooter(
                        dbc.Button("Close", id="close-read-modal", className="ms-auto", n_clicks=0)
                    ),
                ],
                id="read-modal",
                is_open=False,
            ),
            dbc.Modal(
                [
                    dbc.ModalHeader("Operation Status"),
                    dbc.ModalBody(id='initialise-modal-body'),
                    dbc.ModalFooter(
                        dbc.Button("Close", id="close-initialise-modal", className="ms-auto", n_clicks=0)
                    ),
                ],
                id="initialise-modal",
                is_open=False,
            ),
            dbc.Modal(
                [
                    dbc.ModalHeader("Operation Status"),
                    dbc.ModalBody(id='start-modal-body'),
                    dbc.ModalFooter(
                        dbc.Button("Close", id="close-start-modal", className="ms-auto", n_clicks=0)
                    ),
                ],
                id="start-modal",
                is_open=False,
            ),
            dbc.Modal(
                [
                    dbc.ModalHeader("Operation Status"),
                    dbc.ModalBody(id='stop-modal-body'),
                    dbc.ModalFooter(
                        dbc.Button("Close", id="close-stop-modal", className="ms-auto", n_clicks=0)
                    ),
                ],
                id="stop-modal",
                is_open=False,
            ),
            dbc.Modal(
                [
                    dbc.ModalHeader("Operation Status"),
                    dbc.ModalBody(id='reset-modal-body'),
                    dbc.ModalFooter(
                        dbc.Button("Close", id="close-reset-modal", className="ms-auto", n_clicks=0)
                    ),
                ],
                id="reset-modal",
                is_open=False,
            ),
            html.Div(
        id='completion-message',
        className='message-box',  # Apply the general box style
        style={'textAlign': 'center', 'marginTop': '20px', 'fontSize': '20px'}
    ),
    dcc.Interval(
        id='interval-component-data',
        interval=5000,  # Update every 5 seconds
        n_intervals=0
    ),

            dcc.Tabs(id='tabs', value='tab-input', children=[
                dcc.Tab(label='Order List change', value='tab-manage-products', children=[
                    dbc.Row([
                        dbc.Col(
                            dcc.Dropdown(
                                id='manage-dropdown',
                                options=[
                                    {'label': 'Add Order', 'value': 'add'},
                                    {'label': 'Delete Order', 'value': 'delete'}
                                ],
                                value='add',
                                placeholder='Select action',
                                style={'width': '200px'}
                            ),
                            width=3,
                            style={'padding': '20px'}
                        ),
                        dbc.Col(
                            html.Div(id='manage-content'),
                            width=9
                        )
                    ])
                ]),
                
                dcc.Tab(label='Order Catalogue', value='tab-2', children=[
                html.H2('Below are the product details', style={'textAlign': 'left', 'marginBottom': '30px', 'fontSize': '20px'}),
                html.Div([
                    dash_table.DataTable(
                        id='data-table',
                        columns=[],  # Columns will be dynamically generated
                        data=[],  # Data will be dynamically generated
                        filter_action='native',
                        sort_action="native",
                        page_size=10,
                        style_table={'height': '900px', 'overflowY': 'auto', 'marginBottom': '20px'},
                        style_cell={
                            'textAlign': 'center',
                            'padding': '5px',
                            'backgroundColor': '#f9f9f9',
                            'border': '1px solid black',
                            'minWidth': '120px', 'maxWidth': '150px', 'whiteSpace': 'normal'
                        },
                        style_header={
                            'backgroundColor': '#4CAF50',
                            'fontWeight': 'bold',
                            'color': 'white',
                            'border': '1px solid black'
                        }
                    ),
                    dcc.Interval(
                        id='interval-component-table',
                        interval=5000,
                        n_intervals=0
                    )
                ])
            ]),
                
                dcc.Tab(label='Modify', value='tab-modify', children=[
    html.Div([
        html.H2('Modify Order Details by Order Number',
                style={'textAlign': 'left', 'marginBottom': '20px', 'marginTop': '20px', 'fontSize': '20px'}),

        dbc.Row([
            dbc.Col(
                dcc.Input(
                    id='modify-order-number-input',
                    type='number',
                    placeholder='Enter Order Number',
                    style={'marginBottom': '20px', 'width': '200px'}
                ),
                width=3
            ),
            dbc.Col(
                html.Button('Fetch Order', id='fetch-order-button', n_clicks=0, style={'marginBottom': '20px'}),
                width=3
            )
        ]),

        # Table to display order details fetched from the Excel file
        html.Div(id='modify-order-details', style={'marginTop': '20px'}),

        # Buttons for submitting the modification
        dbc.Row([
            dbc.Col(
                html.Button('Save Changes', id='save-modifications-button', n_clicks=0, style={'marginTop': '20px'}),
                width=6
            ),
            dbc.Col(
                html.Div(id='save-confirmation-message', style={'marginTop': '20px'}),
                width=6
            )
        ])
    ])
]),
                
                dcc.Tab(label='Visualize', value='tab-output', children=[
    html.Div(
        "Select a plot to display:",
        style={'textAlign': 'center', 'fontSize': '18px', 'marginTop': '20px'}
    ),
    dcc.Dropdown(
        id='plot-dropdown',
        options=[
            {'label': 'Gantt Chart', 'value': 'Gantt Chart'},
            {'label': 'Operator Utilization', 'value': 'Operator Utilization'},
            {'label': 'Machine Utilization', 'value': 'Machine Utilization'},
            {'label': 'Operator Utilization (Individual)', 'value': 'Operator Utilization (Individual)'},
            {'label': 'Order Status', 'value': 'Order Status'},
            {'label': 'LS_OrderStatus', 'value': 'LS_OrderStatus'},
            {'label': 'P_OrderStatus', 'value': 'P_OrderStatus'}
        ],
        value='Gantt Chart',
        style={'width': '50%', 'margin': '15px auto'}
    ),
    html.Div(
        id='graph-container',
        style={'width': '100%', 'height': '600px'},
        children=[
            dcc.Graph(
                id='main-graph',
                style={'width': '100%', 'height': '100%', 'marginTop': '30px', 'marginBottom': '30px'}
            ),
        ]
    ),
    dcc.Interval(
        id='interval-component-data',
        interval=5000,
        n_intervals=0
    )
]),

                
        # Instructions section
        dcc.Tab(label='Instructions', value='tab-instructions', children=[
            html.Div([
                html.H2('Instructions', style={'marginTop': '30px'}),
                html.Ul([  # Use Unordered List for the points
                    html.Li("The Read from spreadsheet button, checks if the Product Details file is valid and retrieves details"),
                    html.Li("The Start button, starts the scheduling of products with data in Product details file."),
                    html.Li("The Pause button, used to pause the process for making any changes in the data"),
                    html.Li("The Reschedule button is used to continue the scheduling after the changes are made"),
                    html.Li("The Reset button, reset the data to intial status"),
                    html.Li("Use the 'Order List change' tab to add and delete orders in your list."),
                    html.Li("The 'Order Catalogue' tab shows detailed order information."),
                    html.Li("In the 'Modify' tab, you can adjust order data."),
                    html.Li("The 'Visualise' tab contains, Gantt chart, KPI's visualization"),
                    html.Li("The dashboard refreshes automatically. Use the buttons above to Read spreadsheet, start, pause, or restart after updating data.")
                ], style={'fontSize': '16px', 'lineHeight': '1.6em'})
            ])
        ])


            ], style={'width': '100%','marginTop': '50px', 'marginBottom': '50px'}),
        
        ]
    )

    @app.callback(
    Output('completion-message', 'children'),
    Output('completion-message', 'className'),
    Input('interval-component-data', 'n_intervals')
    )
    def update_completion_message(n_intervals):
        # Fetch data from your source
        df1 = fetch_data_table(full_file_path, 'Output')
        #df1 = product_df  # Example data for testing
        if df1 is None or df1.empty:
            raise ValueError("Data is empty or not loaded correctly.")
        statuses = set(df1['Status'])
        if statuses.issubset({'Completed', 'Late'}):
            return "Production scheduling is Completed", 'message-box highlight-green'
        else:
            return "Production Scheduling is InProgress", 'message-box blinking'
        
    

    @app.callback(
        Output('manage-content', 'children'),
        Input('manage-dropdown', 'value')
    )
    def render_manage_content(action):
        if action == 'add':
            return html.Div(
                id='input-form',
                children=[
                    html.H2('Add New Order', style={'textAlign': 'left', 'marginBottom': '30px', 'fontSize': '20px'}),
                    dbc.InputGroup(
                        [
                            dbc.InputGroupText("OrderNumber"),
                            dbc.Input(id='Order-No', type='number', placeholder='Enter Order number'),
                        ],
                        style={'marginBottom': '10px'}
                    ),
                    dbc.InputGroup(
                        [
                            dbc.InputGroupText("Customer Name:"),
                            dbc.Input(id='Customer-Name', type='text', placeholder='Enter Customer name'),
                        ],
                        style={'marginBottom': '10px'}
                    ),
                    dbc.InputGroup(
                        [
                            dbc.InputGroupText("Customer Number:"),
                            dbc.Input(id='Customer-No', type='text', placeholder='Enter Customer Number'),
                        ],
                        style={'marginBottom': '10px'}
                    ),
                    dbc.InputGroup(
                        [
                            dbc.InputGroupText("Order Type:"),
                            dbc.Input(id='Order-Type', type='text', placeholder='Enter Order Type'),
                        ],
                        style={'marginBottom': '10px'}
                    ),
                    dbc.InputGroup(
                        [
                            dbc.InputGroupText("Order Received Date:"),
                            dbc.Input(id='Order-Received-Date', type='date', placeholder='Enter Received date'),
                        ],
                        style={'marginBottom': '10px'}
                    ),
                    dbc.InputGroup(
                        [
                            dbc.InputGroupText("Order Received Time:"),
                            dbc.Input(id='Received-time', type='text', placeholder='Enter Received Time'),
                        ],
                        style={'marginBottom': '10px'}
                    ),
                    dbc.InputGroup(
                        [
                            dbc.InputGroupText("Due Date:"),
                            dbc.Input(id='Promised-Due-Date', type='date', placeholder='Enter Due date'),
                        ],
                        style={'marginBottom': '10px'}
                    ),
                    dbc.InputGroup(
                        [
                            dbc.InputGroupText("Due Time:"),
                            dbc.Input(id='Due-Time', type='text', placeholder='Enter Due Ti'),
                        ],
                        style={'marginBottom': '10px'}
                    ),
                    dbc.InputGroup(
                        [
                            dbc.InputGroupText("Quantity Required:"),
                            dbc.Input(id='Quantity-Required', type='number', placeholder='Enter required quantity'),
                        ],
                        style={'marginBottom': '10px'}
                    ),
                    dbc.InputGroup(
                        [
                            dbc.InputGroupText("Service Type:"),
                            dbc.Input(id='Service-type', type='text', placeholder='Enter Service Type'),
                        ],
                        style={'marginBottom': '10px'}
                    ),
                    dbc.InputGroup(
                        [
                            dbc.InputGroupText("Initial Process Time Per Item:"),
                            dbc.Input(id='inProcess-time', type='number', placeholder='Enter Initial Process Time'),
                        ],
                        style={'marginBottom': '10px'}
                    ),
                    dbc.InputGroup(
                        [
                            dbc.InputGroupText("ExtraService:"),
                            dbc.Input(id='Extra-Service', type='text', placeholder='Enter ExtraService'),
                        ],
                        style={'marginBottom': '10px'}
                    ),
                    dbc.InputGroup(
                                [
                                    dbc.InputGroupText("BagHandling:"),
                                    dcc.Dropdown(
                                        id='Process-Type',
                                        options=[
                                            {'label': 'Basic', 'value': 'Bagging: Basic'},
                                            {'label': 'Advanced', 'value': 'Bagging: Advanced'}
                                        ],
                                        placeholder='Select Bag Handling...',
                                        style={'width': '70%'}  # Adjust width as needed
                                    ),
                                ],
                                style={'marginBottom': '10px'}
                            ),
                    dbc.InputGroup(
                        [
                            dbc.InputGroupText("Baggage/ Unbaggage BeginTime:"),
                            dbc.Input(id='Begin-time', type='text', placeholder='Baggage/ Unbaggage BeginTime (mins)'),
                        ],
                        style={'marginBottom': '10px'}
                    ),
                    dbc.InputGroup(
                        [
                            dbc.InputGroupText("Baggage/ Unbaggage End Time:"),
                            dbc.Input(id='End-time', type='text', placeholder='Enter Baggage/ Unbaggage End Time (mins)'),
                        ],
                        style={'marginBottom': '10px'}
                    ),
                    dbc.InputGroup(
                        [
                            dbc.InputGroupText("IsSecondHand:"),
                            dbc.Input(id='Second-hand', type='number', placeholder='Enter Second hand'),
                        ],
                        style={'marginBottom': '10px'}
                    ),
                    
                    dbc.InputGroup(
                                [
                                    dbc.InputGroupText("MarkingProcess:"),
                                    dcc.Dropdown(
                                        id='MarkingProcess-Type',
                                        options=[
                                            {'label': 'P', 'value': 'P'},
                                            {'label': 'LS', 'value': 'LS'}
                                        ],
                                        placeholder='Select Marking Process...',
                                        style={'width': '70%'}  # Adjust width as needed
                                    ),
                                ],
                                style={'marginBottom': '10px'}
                            ),
                    dbc.InputGroup(
                        [
                            dbc.InputGroupText("Process Time Per Item:"),
                            dbc.Input(id='Process-Time', type='number', placeholder='Enter cycle time'),
                        ],
                        style={'marginBottom': '10px'}
                    ),
                    dbc.InputGroup(
                        [
                            dbc.InputGroupText("Process Type:"),
                            dbc.Input(id='Process-Type', type='text', placeholder='Enter machine numberProcess Type'),
                        ],
                        style={'marginBottom': '10px'}
                    ),
                    html.Button('Submit', id='submit-button', n_clicks=0, style={'marginTop': '10px'}),
                    html.Div(id='submit-output', style={'marginTop': '10px'})
                ]
            )

        elif action == 'delete':
            return html.Div(
                id='input-form',
                children=[
                    html.H2('Delete Product', style={'textAlign': 'left', 'marginBottom': '30px', 'fontSize': '20px'}),
                    dbc.InputGroup(
                        [
                            dbc.InputGroupText("Order Number:"),
                            dbc.Input(id='UniqueID-delete', type='number', placeholder='Enter Order Number to delete'),
                        ],
                        style={'marginBottom': '10px'}
                    ),
                    html.Button('Delete', id='delete-button', n_clicks=0, style={'marginTop': '10px'}),
                    html.Div(id='delete-output', style={'marginTop': '10px'})
                ]
            )
        else:
            return html.Div()

    def modify_new_data(new_data, full_file_path):
        try:
            # Load the Excel file
            print(full_file_path)
            wb = load_workbook(full_file_path)
            sheet = wb["Data"]
            
            # Find the next available row for insertion
            next_row = sheet.max_row + 1
            print(f"Next row{next_row}")
            # Assuming new_data is a dictionary or tuple containing the values
            sheet.append(new_data)
            
            # Save the changes
            wb.save(full_file_path)
            wb.close()
            
            return True
        
        except Exception as e:
            print(f"Error modifying data: {e}")
            return False
        
    # Callback to add a new order to the Excel file
    @app.callback(
        Output('submit-output', 'children'),
        Input('submit-button', 'n_clicks'),
        State('Order-No', 'value'),
        State('Customer-Name', 'value'),
        State('Customer-No', 'value'),
        State('Order-Type', 'value'),
        State('Order-Received-Date', 'value'),
        State('Received-time', 'value'),
        State('Promised-Due-Date', 'value'),
        State('Due-Time', 'value'),
        State('Quantity-Required', 'value'),
        State('Service-type', 'value'),
        State('inProcess-time', 'value'),
        State('Extra-Service', 'value'),
        State('Begin-time', 'value'),
        State('End-time', 'value'),
        State('Second-hand', 'value'),
        State('MarkingProcess-Type', 'value'),
        State('Process-Time', 'value'),
        State('Process-Type', 'value')
    )
    def add_new_order(n_clicks, order_no, customer_name, customer_no, order_type, received_date, received_time, due_date, due_time,
                    quantity, service_type, process_time_per_item, extra_service, baggage_start, baggage_end, second_hand, 
                    marking_process, cycle_time, process_type):
        if n_clicks > 0:
            try:
                # Load the workbook and get the Output sheet
                wb = load_workbook(full_file_path)
                sheet = wb["Output"]

                # Prepare the new data to append
                new_data = [order_no, customer_name, customer_no, order_type, received_date, received_time, due_date, due_time, 
                            quantity, service_type, process_time_per_item, extra_service, None, baggage_start, baggage_end,
                            second_hand, marking_process, cycle_time, process_type, None, "InProgress", None, None, None, None, None, None]

                # Append the new data to the sheet
                sheet.append(new_data)

                # Save and close the workbook
                wb.save(full_file_path)
                wb.close()

                return dbc.Alert("Order added successfully", color="success", dismissable=True)
            except Exception as e:
                return dbc.Alert(f"Error adding order: {e}", color="danger", dismissable=True)
        return ""


    # Callback to delete an order from the Excel file
    @app.callback(
        Output('delete-output', 'children'),
        Input('delete-button', 'n_clicks'),
        State('UniqueID-delete', 'value')
    )
    def delete_order(n_clicks, order_no):
        if n_clicks > 0:
            try:
                # Load the workbook and get the Output sheet
                wb = load_workbook(full_file_path)
                sheet = wb["Output"]

                # Find the row corresponding to the order number and delete it
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=False):
                    if row[0].value == order_no:  # Assuming order number is in the first column
                        sheet.delete_rows(row[0].row)
                        break

                # Save and close the workbook
                wb.save(full_file_path)
                wb.close()

                return dbc.Alert("Order deleted successfully", color="success", dismissable=True)
            except Exception as e:
                return dbc.Alert(f"Error deleting order: {e}", color="danger", dismissable=True)
        return ""

        
    @app.callback(
    Output('data-table', 'columns'),
    Output('data-table', 'data'),
    Input('interval-component-table', 'n_intervals')
    )
    def update_order_catalogue(n_intervals):
        try:
            df = fetch_data_table(full_file_path, 'Output')
            if df is not None:
                columns = [{'name': col, 'id': col} for col in df.columns]
                data = df.to_dict('records')
                return columns, data
        except Exception as e:
            return [], []
        return [], []

    @app.callback(
    Output('modify-order-details', 'children'),
    Input('fetch-order-button', 'n_clicks'),
    State('modify-order-number-input', 'value')
    )
    def fetch_order_details(n_clicks, order_number):
        if n_clicks > 0 and order_number is not None:
            try:
                # Load the Excel workbook and find the order in the 'Output' sheet
                wb = load_workbook(full_file_path)
                sheet = wb['Output']
                order_data = None

                # Find the row corresponding to the order number
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
                    if row[0] == order_number:
                        order_data = row
                        break
                
                wb.close()

                if order_data:
                    # Display the details in an editable form
                    columns = ['OrderNumber', 'CustomerName', 'Customer Number', 'Order Type', 'ReceivedDate', 'ReceivedTime',
                            'DueDate', 'DueTime', 'Quantity', 'Service Type', 'Initial Process Time Per Item', 'ExtraService',
                            'BagHandling', 'Baggage/ Unbaggage BeginTime', 'Baggage/ Unbaggage', 'End Time', 'IsSecondHand',
                            'MarkingProcess', 'Process Time Per Item', 'Process Type', 'Pass or Fail', 'Status']
                    
                    order_details = [
                        dbc.InputGroup([
                            dbc.InputGroupText(col),
                            dbc.Input(id={'type': 'order-detail-input', 'index': i}, value=order_data[i], type='text')
                        ], style={'marginBottom': '10px'}) for i, col in enumerate(columns)
                    ]

                    return html.Div(order_details)
                else:
                    return dbc.Alert("Order not found.", color="danger")
            except Exception as e:
                return dbc.Alert(f"Error fetching order: {e}", color="danger")
        
        return ""

    @app.callback(
    Output('save-confirmation-message', 'children'),
    Input('save-modifications-button', 'n_clicks'),
    State('modify-order-number-input', 'value'),
    State({'type': 'order-detail-input', 'index': ALL}, 'value')
    )
    def save_modified_order(n_clicks, order_number, modified_values):
        if n_clicks > 0 and order_number is not None and modified_values:
            try:
                # Load the Excel workbook and find the order in the 'Output' sheet
                wb = load_workbook(full_file_path)
                sheet = wb['Output']

                # Find the row corresponding to the order number
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=False):
                    if row[0].value == order_number:
                        # Update the row with modified values
                        for i, value in enumerate(modified_values):
                            row[i].value = value
                        break

                # Save and close the workbook
                wb.save(full_file_path)
                wb.close()

                return dbc.Alert("Order updated successfully!", color="success")
            except Exception as e:
                return dbc.Alert(f"Error updating order: {e}", color="danger")

        return ""






    def read_output(process):
        try:
            for stdout_line in iter(process.stdout.readline, ''):
                if stdout_line:
                    print(f"STDOUT: {stdout_line.strip()}")
            process.stdout.close()
        except Exception as e:
            print(f"Error reading stdout: {e}")

        try:
            for stderr_line in iter(process.stderr.readline, ''):
                if stderr_line:
                    print(f"STDERR: {stderr_line.strip()}")
            process.stderr.close()
        except Exception as e:
            print(f"Error reading stderr: {e}")


    def start_allocation_check(keyword):
        global allocation_process
        # Print statements to debug
        print("Starting BOM.py...")
        print(f"Keyword: {keyword}")
        
        # Verifying if the file exists
        script_path = r'BOM.py'
        if not os.path.exists(script_path):
            print(f"Script not found at path: {script_path}")
            return
        try:
        
            allocation_process = subprocess.Popen(
                [sys.executable, script_path],
                stdin=subprocess.PIPE,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                bufsize=1  # Line-buffered mode
            )
            # Write the keyword to the process and flush
            print(f"Sending keyword to BOM.py: {keyword}")
            allocation_process.stdin.write(keyword + '\n')
            allocation_process.stdin.flush()
            #stdout, stderr = allocation_process.communicate(keyword)
            # Start a thread to read the output and errors asynchronously
            threading.Thread(target=read_output, args=(allocation_process,), daemon=True).start()
            print("BOM.py started successfully")
        except Exception as e:
            print(f"Error starting BOM.py: {e}")
            raise
            


    # Function to stop the execution of the external script
    def stop_allocation_check():
        global allocation_process
        if allocation_process and allocation_process.poll() is None:
            allocation_process.terminate()
            allocation_process = None
    
    @app.callback(
        [
            Output('interval-component-script', 'disabled'),
            Output('start-message', 'children'),
            Output('read-modal', 'is_open'),
            Output('initialise-modal', 'is_open'),
            Output('start-modal', 'is_open'),
            Output('stop-modal', 'is_open'),
            Output('reset-modal', 'is_open'),
            Output('read-modal-body', 'children'),
            Output('initialise-modal-body', 'children'),
            Output('start-modal-body', 'children'),
            Output('stop-modal-body', 'children'),
            Output('reset-modal-body', 'children')
        ],
        [
            Input('read-button', 'n_clicks'),
            Input('initialise-button', 'n_clicks'),
            Input('start-button', 'n_clicks'),
            Input('stop-button', 'n_clicks'),
            Input('reset-button', 'n_clicks'),
            Input('close-read-modal', 'n_clicks'),
            Input('close-initialise-modal', 'n_clicks'),
            Input('close-start-modal', 'n_clicks'),
            Input('close-stop-modal', 'n_clicks'),
            Input('close-reset-modal', 'n_clicks')
        ],
        [State('interval-component-script', 'disabled')]
    )
    def control_allocation_check(read_clicks, initialise_clicks, start_clicks, stop_clicks, reset_clicks, close_read_modal,
                                close_initialise_modal, close_start_modal, close_stop_modal, close_reset_modal,
                                interval_disabled):
        ctx = dash.callback_context
        if not ctx.triggered:
            return interval_disabled, None, False, False, False, False, False, "", "", "", "", ""
        
        triggered_id = ctx.triggered[0]['prop_id'].split('.')[0]
        read_modal_open=False
        initialise_modal_open = False
        start_modal_open = False
        stop_modal_open = False
        reset_modal_open = False
        read_modal_body = ""
        initialise_modal_body = ""
        start_modal_body = ""
        stop_modal_body = ""
        modal_body = "Process started..."

        if triggered_id == 'read-button' and read_clicks:
            # Open the modal immediately
            read_modal_open = True
           
            if os.path.exists(full_file_path):
                print("File Exists")
                modal_body = "Data is retrieved from spreadsheet. Click on the Start button to start scheduling process"
            else:
                print("File Not found, Please check")
                modal_body = "Product Details File Not found, Please check"
            # Start the process asynchronously
            #threading.Thread(target=start_allocation_check, args=("Initial",), daemon=True).start()
            interval_disabled = False
            
        elif triggered_id == 'initialise-button' and initialise_clicks:
            # Open the modal immediately
            initialise_modal_open = True
            modal_body = "Scheduling Process Started."
            # Start the process asynchronously
            threading.Thread(target=start_allocation_check, args=("Initial",), daemon=True).start()
            interval_disabled = False
 
        elif triggered_id == 'start-button' and start_clicks:
            start_modal_open = True
            modal_body = "Scheduling Process Started again."
            threading.Thread(target=start_allocation_check, args=("Start",), daemon=True).start()
            interval_disabled = False

        elif triggered_id == 'stop-button' and stop_clicks:
            stop_modal_open = True
            modal_body = "Program paused, Make the necessary changes, If any. and click on the Reschedule button."
            stop_allocation_check()
            interval_disabled = True
        elif triggered_id == 'reset-button' and reset_clicks:
            reset_modal_open = True
            modal_body = "Product Details have been reset."

            # Perform the reset functionality
            if os.path.exists(full_file_path):
                try:
                    # Read the 'P' sheet
                    df_p = pd.read_excel(full_file_path, sheet_name='Data')
                    
                    # Open the workbook and delete the 'Output' sheet if it exists
                    with pd.ExcelWriter(full_file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                        # Load the existing workbook
                        book = writer.book
                        
                        # Check if the 'Output' sheet exists
                        if 'Output' in book.sheetnames:
                            # Remove the 'Output' sheet
                            std = book['Output']
                            book.remove(std)
                        
                        # Write the updated data to the 'Output' sheet
                        df_p.to_excel(writer, sheet_name='Output', index=False)
                        
                    modal_body += " Data copied from 'Data' sheet to 'Output' sheet successfully."
                except Exception as e:
                    modal_body = f"An error occurred: {e}"
            else:
                modal_body = "Product Details File Not found, Please check"
            
            interval_disabled = True

        elif triggered_id.startswith('close'):
            return interval_disabled, None, False, False, False, False, False, "", "", "","",""
        
        return (interval_disabled, None,read_modal_open, initialise_modal_open, start_modal_open, stop_modal_open, reset_modal_open,
                modal_body if read_modal_open else "",
                modal_body if initialise_modal_open else "",
                modal_body if start_modal_open else "",
                modal_body if stop_modal_open else "",
                modal_body if reset_modal_open else "")


            
   
    @app.callback(
    Output('main-graph', 'figure'),
    [Input('plot-dropdown', 'value'), Input('interval-component-data', 'n_intervals')]
    )
    def update_graph(selected_plot, n):
        #df = fetch_data()
        
        data=pd.read_excel("Data Sample.xlsx", sheet_name='Output')
        
        if data.empty:
            return px.line(title='No Data Available')

        # Process the DataFrame for each plot type
        if selected_plot=="Gantt Chart":
           
            # Convert StartTime and EndTime to datetime
            data['StartTime'] = pd.to_datetime(data['StartTime'])
            data['EndTime'] = pd.to_datetime(data['EndTime'])

            # Set a standard setup time of 5 minutes for all orders
            setup_time = timedelta(minutes=5)

            # Calculate SetupStartTime (StartTime - 5 minutes)
            data['SetupStartTime'] = data['StartTime'] - setup_time

            # Add a column for setup time (5 minutes for each order)
            data['SetupDuration'] = '5 minutes'

            # Format the StartTime, EndTime, and SetupStartTime
            data['Formatted_StartTime'] = data['StartTime'].dt.strftime('%d-%b %H:%M')
            data['Formatted_EndTime'] = data['EndTime'].dt.strftime('%d-%b %H:%M')
            data['Formatted_SetupStartTime'] = data['SetupStartTime'].dt.strftime('%d-%b %H:%M')

            # Dynamically set the height of the chart based on the number of orders
            chart_height = len(data) * 50

            # Get the minimum SetupStartTime to set as the start of the x-axis
            min_start_time = data['SetupStartTime'].min()

            # Create Gantt Chart
            fig = px.timeline(
                data,
                x_start="StartTime",
                x_end="EndTime",
                y="Order Type",
                color='Status',
                custom_data=[data['Formatted_StartTime'], data['Formatted_EndTime'], data['MarkingProcess'], data['Process Type'], data['Used_Machines'], data['Normal_Operators_Used'], data['High_Operators_Used']],
                title="Gantt Chart with Order Details",
                color_discrete_map={'Late': 'red', 'Completed': 'green'}
            )

            # Add a trace for the Setup Time
            setup_trace = px.timeline(
                data,
                x_start="SetupStartTime",
                x_end="StartTime",
                y="Order Type",
                color_discrete_sequence=['black']
            ).data[0]

            # Modify the setup trace to appear in the legend with the label "Setup Time"
            setup_trace.update(
                showlegend=True,
                name="Setup Time"
            )

            # Add the modified setup trace to the figure
            fig.add_trace(setup_trace)

            # Add a dummy trace for "Late" status with square markers
            fig.add_trace(go.Scatter(
                x=[None], y=[None],
                mode='markers',
                marker=dict(color='red', symbol='square', size=10),  # Square marker
                name='Late'
            ))

            # Customize the hover template
            fig.update_traces(hovertemplate=
                'Order: %{y}<br>' +
                'Start Date: %{customdata[0]}<br>' +
                'End Date: %{customdata[1]}<br>' +
                'Marking Process: %{customdata[2]}<br>' +
                'Process Type: %{customdata[3]}<br>' +
                'Machines: %{customdata[4]}<br>' +
                'Normal Skilled Operators: %{customdata[5]}<br>' +
                'High Skilled Operators: %{customdata[6]}<br>'
            )

            # Set X-axis ticks to 30-minute intervals
            fig.update_xaxes(
                tickformat="%d-%b %H:%M",
                dtick=30*60*1000,  # 30 minutes in milliseconds
                ticklabelmode="period",
                range=[min_start_time, data['EndTime'].max()],
                tickfont=dict(size=16, weight='bold')
            )

            # Preserve the order of the orders as they appear in the original file
            fig.update_layout(
                xaxis_title="Time",
                yaxis_title="Order",
                title="Gantt Chart of Order Processing (With Setup Time and Date)",
                height=chart_height,
                yaxis=dict(
                    categoryarray=data['Order Type'],
                    categoryorder="array",
                    tickfont=dict(size=16, weight='bold')  # Increase font size and make it bold
                ),
                margin=dict(l=150, r=50, b=50, t=50),
                showlegend=True
            )

            # Set bold titles for x-axis and y-axis using weight instead of bold
            fig.update_xaxes(title_font=dict(size=14, color='black', family='Arial', weight='bold'))
            fig.update_yaxes(title_font=dict(size=14, color='black', family='Arial', weight='bold'))

    
        elif selected_plot == "Operator Utilization":
            normal_operator_usage = data[data['MarkingProcess'] == 'P']['Normal_Operators_Used'].sum()
            high_operator_usage = data[data['MarkingProcess'] == 'P']['High_Operators_Used'].sum()

            # Create a bar chart for operator utilization using Plotly
            fig = go.Figure(data=[
                go.Bar(name='Normal-skilled', x=['Normal-skilled'], y=[normal_operator_usage], marker_color='orange'),
                go.Bar(name='High-skilled', x=['High-skilled'], y=[high_operator_usage], marker_color='red')
            ])

            # Update the layout
            fig.update_layout(
                title='Operator Utilization for P-type Process',
                xaxis_title='Operator Type',
                yaxis_title='Number of Operators Used',
                barmode='group',
                height=500
            )

        elif selected_plot == "Machine Utilization":
           # Split machine usage for P and LS
            # Split machine usage for P and LS
            p_machines_usage = data[data['MarkingProcess'] == 'P']['Used_Machines'].str.split(',', expand=True).stack().value_counts()
            ls_machines_usage = data[data['MarkingProcess'] == 'LS']['Used_Machines'].str.split(',', expand=True).stack().value_counts()

            # Calculate total orders for percentage calculation
            total_p_orders = data[data['MarkingProcess'] == 'P'].shape[0]
            total_ls_orders = data[data['MarkingProcess'] == 'LS'].shape[0]

            # Calculate percentage usage
            p_machines_percentage = (p_machines_usage / total_p_orders) * 100
            ls_machines_percentage = (ls_machines_usage / total_ls_orders) * 100

            # Create bar charts for both P-type and LS-type processes
            fig = make_subplots(rows=1, cols=2, subplot_titles=('P-type Machine Utilization (%)', 'LS-type Machine Utilization (%)'))

            # Plot Machine Utilization for P-type process
            fig.add_trace(
                go.Bar(x=p_machines_usage.index, y=p_machines_usage.values, marker_color='blue', name="P-type Machines"),
                row=1, col=1
            )

            # Plot Machine Utilization for LS-type process
            fig.add_trace(
                go.Bar(x=ls_machines_usage.index, y=ls_machines_usage.values, marker_color='green', name="LS-type Machines"),
                row=1, col=2
            )

            # Update the layout for both subplots
            fig.update_layout(
                title_text='Machine Utilization for P-type and LS-type Processes',
                showlegend=False,
                height=500
            )

            # Update axis titles
            fig.update_xaxes(title_text="Machine", row=1, col=1)
            fig.update_yaxes(title_text="Number of Orders Processed", row=1, col=1, tickvals=[0, 20, 40, 60, 80, 100])
            fig.update_xaxes(title_text="Machine", row=1, col=2)
            fig.update_yaxes(title_text="Number of Orders Processed", row=1, col=2, tickvals=[0, 1, 2, 3, 4, 5])

        elif selected_plot=="Operator Utilization (Individual)":
            # Split machine usage for P and LS
            p_machines_usage = data[data['MarkingProcess'] == 'P']['Used_Operators'].str.split(',', expand=True).stack().value_counts()
            ls_machines_usage = data[data['MarkingProcess'] == 'LS']['Used_Operators'].str.split(',', expand=True).stack().value_counts()


            # Calculate total orders for percentage calculation
            total_p_orders = data[data['MarkingProcess'] == 'P'].shape[0]
            total_ls_orders = data[data['MarkingProcess'] == 'LS'].shape[0]

            # Calculate percentage usage
            p_machines_percentage = (p_machines_usage / total_p_orders) * 100
            ls_machines_percentage = (ls_machines_usage / total_ls_orders) * 100

            # Create bar charts for both P-type and LS-type processes
            fig = make_subplots(rows=1, cols=2, subplot_titles=('P-type Operator Utilization (%)', 'LS-type Operator Utilization (%)'))

            # Plot Machine Utilization for P-type process
            fig.add_trace(
                go.Bar(x=p_machines_usage.index, y=p_machines_usage.values, marker_color='blue', name="P-type Operators"),
                row=1, col=1
            )

            # Plot Machine Utilization for LS-type process
            fig.add_trace(
                go.Bar(x=ls_machines_usage.index, y=ls_machines_usage.values, marker_color='green', name="LS-type Operators"),
                row=1, col=2
            )

            # Update the layout for both subplots
            fig.update_layout(
                title_text='Operators Utilization for P-type and LS-type Processes',
                showlegend=False,
                height=500
            )

            # Update axis titles
            fig.update_xaxes(title_text="Operators", row=1, col=1)
            fig.update_yaxes(title_text="Number of Orders Processed", row=1, col=1, tickvals=[0, 20, 40, 60, 80, 100])
            fig.update_xaxes(title_text="Operators", row=1, col=2)
            fig.update_yaxes(title_text="Number of Orders Processed", row=1, col=2, tickvals=[0, 1, 2, 3, 4, 5])
        elif selected_plot=="Order Status":
            # Clean up the Status column
            data['Status'] = data['Status'].apply(lambda x: str(x).strip() if x is not None else '')

            # Define colors for the different statuses
            status_colors = {
                "Completed": "green",
                "InProgress": "yellow",
                "Late": "red"
            }

            # Generate the figure
            fig = go.Figure()

            # Maximum number of orders per row
            orders_per_row = 5
            order_counter = 0

            # Loop through each order and add them to the plot
            for index, row in data.iterrows():
                order_number = row['OrderNumber']
                status = row['Status']
                marking_process = row['MarkingProcess']
                process_type = row['Process Type']
                used_machines = row['Used_Machines']
                normal_operators = row['Normal_Operators_Used']
                high_operators = row['High_Operators_Used']
                
                start_date = pd.to_datetime(row['StartTime'], errors='coerce')
                end_date = pd.to_datetime(row['EndTime'], errors='coerce')
                
                # Set the color based on the status
                color = status_colors.get(status, "grey")
                
                # Calculate x and y position to fit 5 orders per row
                x = order_counter % orders_per_row
                y = (order_counter // orders_per_row) + 1  # Positive y-axis rows
                
                # Hover text content
                hover_text = (
                    f"Order: {row['Order Type']}<br>"
                    f"Start Date: {start_date.strftime('%d-%b %H:%M') if pd.notnull(start_date) else 'N/A'}<br>"
                    f"End Date: {end_date.strftime('%d-%b %H:%M') if pd.notnull(end_date) else 'N/A'}<br>"
                    f"Marking Process: {marking_process}<br>"
                    f"Process Type: {process_type}<br>"
                    f"Machines: {used_machines}<br>"
                    f"Normal Skilled Operators: {normal_operators}<br>"
                    f"High Skilled Operators: {high_operators}"
                )
                
                # Add a square marker for each order
                fig.add_trace(go.Scatter(
                    x=[x], y=[y],
                    mode='markers+text',
                    marker=dict(
                        color=color,
                        size=50,
                        symbol='square'
                    ),
                    text=[str(order_number)],  # Display order number inside the block
                    textposition='middle center',
                    hovertemplate=hover_text,  # Add hover information
                    showlegend=False  # Disable individual order markers from showing in the legend
                ))
                
                order_counter += 1

            # Add dummy traces for legend
            fig.add_trace(go.Scatter(
                x=[None], y=[None],
                mode='markers',
                marker=dict(color='yellow', size=10, symbol='square'),
                legendgroup='InProgress',
                showlegend=True,
                name='InProgress'
            ))

            fig.add_trace(go.Scatter(
                x=[None], y=[None],
                mode='markers',
                marker=dict(color='green', size=10, symbol='square'),
                legendgroup='Completed',
                showlegend=True,
                name='Completed'
            ))

            fig.add_trace(go.Scatter(
                x=[None], y=[None],
                mode='markers',
                marker=dict(color='red', size=10, symbol='square'),
                legendgroup='Late',
                showlegend=True,
                name='Late'
            ))

            # Adjust the layout to fit the grid-style view
            fig.update_layout(
                title='Order Status Tracking',
                xaxis=dict(
                    title='Order Blocks (5 per row)',
                    tickvals=list(range(orders_per_row)),  # Set tick values for 5 blocks per row
                    range=[-0.5, orders_per_row - 0.5],
                    showgrid=False
                ),
                yaxis=dict(
                    title='Row of Orders',
                    tickvals=list(range(1, (order_counter // orders_per_row) + 2)),  # Positive y-axis values
                    showgrid=False
                ),
                height=600,  # Adjust height based on the number of rows
                showlegend=True
            )
        elif selected_plot=="P_OrderStatus":
           
                        # Clean up the Status column
            data['Status'] = data['Status'].apply(lambda x: str(x).strip() if x is not None else '')

            # Filter for Marking Process "P"
            data_filtered = data[data['MarkingProcess'] == 'P']

            # Define colors for the different statuses
            status_colors = {
                "Completed": "green",
                "InProgress": "yellow",
                "Late": "red"
            }

            # Generate the figure
            fig = go.Figure()

            # Maximum number of orders per row
            orders_per_row = 5
            order_counter = 0

            # Loop through each filtered order and add them to the plot
            for index, row in data_filtered.iterrows():
                order_number = row['OrderNumber']
                status = row['Status']
                marking_process = row['MarkingProcess']
                process_type = row['Process Type']
                used_machines = row['Used_Machines']
                
                # Used Operators list
                used_operators = row['Used_Operators']
                
                # Start and End dates for the orders
                start_date = pd.to_datetime(row['StartTime'], errors='coerce')
                end_date = pd.to_datetime(row['EndTime'], errors='coerce')
                
                # Set the color based on the status
                color = status_colors.get(status, "grey")
                
                # Calculate x and y position to fit 5 orders per row
                x = order_counter % orders_per_row
                y = (order_counter // orders_per_row) + 1  # Positive y-axis rows
                
                # Create hover text with operator names included
                hover_text = (
                    f"Order: {row['OrderNumber']}<br>"
                    f"Start Date: {start_date.strftime('%d-%b %H:%M') if pd.notnull(start_date) else 'N/A'}<br>"
                    f"End Date: {end_date.strftime('%d-%b %H:%M') if pd.notnull(end_date) else 'N/A'}<br>"
                    f"Marking Process: {marking_process}<br>"
                    f"Process Type: {process_type}<br>"
                    f"Machines: {used_machines}<br>"
                    f"Operators: {used_operators}"
                )
                
                # Add a square marker for each order
                fig.add_trace(go.Scatter(
                    x=[x], y=[y],
                    mode='markers+text',
                    marker=dict(
                        color=color,
                        size=50,
                        symbol='square'
                    ),
                    text=[str(order_number)],  # Display order number inside the block
                    textposition='middle center',
                    hovertemplate=hover_text,  # Add hover information with operator names
                    showlegend=False  # Disable individual order markers from showing in the legend
                ))
                
                order_counter += 1

            # Add dummy traces for legend
            fig.add_trace(go.Scatter(
                x=[None], y=[None],
                mode='markers',
                marker=dict(color='yellow', size=10, symbol='square'),
                legendgroup='InProgress',
                showlegend=True,
                name='InProgress'
            ))

            fig.add_trace(go.Scatter(
                x=[None], y=[None],
                mode='markers',
                marker=dict(color='green', size=10, symbol='square'),
                legendgroup='Completed',
                showlegend=True,
                name='Completed'
            ))

            fig.add_trace(go.Scatter(
                x=[None], y=[None],
                mode='markers',
                marker=dict(color='red', size=10, symbol='square'),
                legendgroup='Late',
                showlegend=True,
                name='Late'
            ))

            # Adjust the layout to fit the grid-style view
            fig.update_layout(
                title='Order Status Tracking for Marking Process "P"',
                xaxis=dict(
                    title='Order Blocks (5 per row)',
                    tickvals=list(range(orders_per_row)),  # Set tick values for 5 blocks per row
                    range=[-0.5, orders_per_row - 0.5],
                    showgrid=False
                ),
                yaxis=dict(
                    title='Row of Orders',
                    tickvals=list(range(1, (order_counter // orders_per_row) + 2)),  # Positive y-axis values
                    showgrid=False
                ),
                height=600,  # Adjust height based on the number of rows
                showlegend=True
            )

        elif selected_plot=="LS_OrderStatus":
            # Clean up the Status column
            data['Status'] = data['Status'].apply(lambda x: str(x).strip() if x is not None else '')

            # Filter for Marking Process "P"
            data_filtered = data[data['MarkingProcess'] == 'LS']

            # Define colors for the different statuses
            status_colors = {
                "Completed": "green",
                "InProgress": "yellow",
                "Late": "red"
            }

            # Generate the figure
            fig = go.Figure()

            # Maximum number of orders per row
            orders_per_row = 5
            order_counter = 0

            # Loop through each filtered order and add them to the plot
            for index, row in data_filtered.iterrows():
                order_number = row['OrderNumber']
                status = row['Status']
                marking_process = row['MarkingProcess']
                process_type = row['Process Type']
                used_machines = row['Used_Machines']
                
                # Used Operators list
                used_operators = row['Used_Operators']
                
                # Start and End dates for the orders
                start_date = pd.to_datetime(row['StartTime'], errors='coerce')
                end_date = pd.to_datetime(row['EndTime'], errors='coerce')
                
                # Set the color based on the status
                color = status_colors.get(status, "grey")
                
                # Calculate x and y position to fit 5 orders per row
                x = order_counter % orders_per_row
                y = (order_counter // orders_per_row) + 1  # Positive y-axis rows
                
                # Create hover text with operator names included
                hover_text = (
                    f"Order: {row['OrderNumber']}<br>"
                    f"Start Date: {start_date.strftime('%d-%b %H:%M') if pd.notnull(start_date) else 'N/A'}<br>"
                    f"End Date: {end_date.strftime('%d-%b %H:%M') if pd.notnull(end_date) else 'N/A'}<br>"
                    f"Marking Process: {marking_process}<br>"
                    f"Process Type: {process_type}<br>"
                    f"Machines: {used_machines}<br>"
                    f"Operators: {used_operators}"
                )
                
                # Add a square marker for each order
                fig.add_trace(go.Scatter(
                    x=[x], y=[y],
                    mode='markers+text',
                    marker=dict(
                        color=color,
                        size=50,
                        symbol='square'
                    ),
                    text=[str(order_number)],  # Display order number inside the block
                    textposition='middle center',
                    hovertemplate=hover_text,  # Add hover information with operator names
                    showlegend=False  # Disable individual order markers from showing in the legend
                ))
                
                order_counter += 1

            # Add dummy traces for legend
            fig.add_trace(go.Scatter(
                x=[None], y=[None],
                mode='markers',
                marker=dict(color='yellow', size=10, symbol='square'),
                legendgroup='InProgress',
                showlegend=True,
                name='InProgress'
            ))

            fig.add_trace(go.Scatter(
                x=[None], y=[None],
                mode='markers',
                marker=dict(color='green', size=10, symbol='square'),
                legendgroup='Completed',
                showlegend=True,
                name='Completed'
            ))

            fig.add_trace(go.Scatter(
                x=[None], y=[None],
                mode='markers',
                marker=dict(color='red', size=10, symbol='square'),
                legendgroup='Late',
                showlegend=True,
                name='Late'
            ))

            # Adjust the layout to fit the grid-style view
            fig.update_layout(
                title='Order Status Tracking for Marking Process "LS"',
                xaxis=dict(
                    title='Order Blocks (5 per row)',
                    tickvals=list(range(orders_per_row)),  # Set tick values for 5 blocks per row
                    range=[-0.5, orders_per_row - 0.5],
                    showgrid=False
                ),
                yaxis=dict(
                    title='Row of Orders',
                    tickvals=list(range(1, (order_counter // orders_per_row) + 2)),  # Positive y-axis values
                    showgrid=False
                ),
                height=600,  # Adjust height based on the number of rows
                showlegend=True
            )

        return fig

    app.run_server(debug=False)


if __name__ == '__main__':
    launch_dashboard()
    
    
    
